from flask import Flask, request, send_file, jsonify, abort
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re, sqlite3, json, os, uuid
from datetime import datetime

app = Flask(__name__, static_folder="../frontend/dist", static_url_path="/")
CORS(app)

DB_PATH = os.environ.get("DB_PATH",
          os.path.join(os.path.dirname(__file__), "analyses.db"))

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS analyses (
            id         TEXT PRIMARY KEY,
            name       TEXT NOT NULL,
            saved_at   TEXT NOT NULL,
            is_draft   INTEGER NOT NULL DEFAULT 0,
            data       TEXT NOT NULL,
            folder_id  TEXT,
            version    INTEGER NOT NULL DEFAULT 1
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS folders (
            id         TEXT PRIMARY KEY,
            name       TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        )
    """)
    for sql in [
        "ALTER TABLE analyses ADD COLUMN is_draft INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE analyses ADD COLUMN folder_id TEXT",
        "ALTER TABLE analyses ADD COLUMN version INTEGER NOT NULL DEFAULT 1",
    ]:
        try:
            conn.execute(sql)
        except Exception:
            pass
    conn.commit()
    conn.close()

init_db()


def _apply_initials_suffix(name, initials):
    base = (name or "").strip()
    ini = (initials or "").strip().upper()
    if not base:
        return base
    if not ini:
        return base
    suffix = f" - {ini}"
    return base if base.endswith(suffix) else base + suffix

@app.route("/api/analyses", methods=["GET"])
def list_analyses():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, name, saved_at, is_draft, folder_id, version FROM analyses ORDER BY saved_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([{"id": r["id"], "name": r["name"],
                     "savedAt": r["saved_at"], "isDraft": bool(r["is_draft"]),
                     "folderId": r["folder_id"], "version": r["version"]}
                    for r in rows])


@app.route("/api/folders", methods=["GET"])
def list_folders():
    conn = get_db()
    rows = conn.execute("SELECT id, name, created_at FROM folders ORDER BY name COLLATE NOCASE").fetchall()
    conn.close()
    return jsonify([{"id": r["id"], "name": r["name"], "createdAt": r["created_at"]} for r in rows])

@app.route("/api/folders", methods=["POST"])
def create_folder():
    body = request.get_json() or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    fid = uuid.uuid4().hex
    now = datetime.now().strftime("%b %d, %I:%M %p")
    conn = get_db()
    try:
        conn.execute("INSERT INTO folders (id, name, created_at) VALUES (?, ?, ?)", (fid, name, now))
        conn.commit()
    except sqlite3.IntegrityError:
        row = conn.execute("SELECT id, name, created_at FROM folders WHERE lower(name)=lower(?)", (name,)).fetchone()
        conn.close()
        return jsonify({"id": row["id"], "name": row["name"], "createdAt": row["created_at"]})
    conn.close()
    return jsonify({"id": fid, "name": name, "createdAt": now})

@app.route("/api/analyses", methods=["POST"])
def save_analysis():
    body     = request.get_json() or {}
    aid      = body.get("id", "").strip()
    raw_name = body.get("name", "").strip()
    data     = body.get("data", {})
    initials = (body.get("initials") or data.get("initials") or "").strip().upper()
    is_draft = 1 if body.get("isDraft") else 0
    folder_id = body.get("folderId")
    expected_version = body.get("expectedVersion")
    force = bool(body.get("force"))
    if not aid or not raw_name:
        return jsonify({"error": "id and name are required"}), 400

    name = _apply_initials_suffix(raw_name, initials) if not is_draft else raw_name
    now = datetime.now().strftime("%b %d, %I:%M %p")
    conn = get_db()
    row = conn.execute("SELECT version, saved_at, is_draft, folder_id FROM analyses WHERE id=?", (aid,)).fetchone()
    current_version = row["version"] if row else 0
    if row and expected_version is not None and int(expected_version) != int(current_version) and not force:
        conn.close()
        return jsonify({"error": "version_conflict", "currentVersion": current_version, "savedAt": row["saved_at"]}), 409

    if row and not row["is_draft"] and is_draft:
        is_draft = 0
        name = row["folder_id"] and name or name

    effective_folder_id = folder_id if folder_id is not None else (row["folder_id"] if row else None)
    new_version = current_version + 1 if row else 1
    conn.execute("""
        INSERT INTO analyses (id, name, saved_at, is_draft, data, folder_id, version)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, saved_at=excluded.saved_at,
            is_draft=excluded.is_draft, data=excluded.data,
            folder_id=excluded.folder_id, version=excluded.version
    """, (aid, name, now, is_draft, json.dumps(data), effective_folder_id, new_version))
    conn.commit()
    conn.close()
    return jsonify({"id": aid, "name": name, "savedAt": now, "isDraft": bool(is_draft), "folderId": effective_folder_id, "version": new_version})

@app.route("/api/analyses/<aid>", methods=["GET"])
def load_analysis(aid):
    conn = get_db()
    row  = conn.execute("SELECT * FROM analyses WHERE id=?", (aid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"id": row["id"], "name": row["name"], "savedAt": row["saved_at"],
                    "isDraft": bool(row["is_draft"]), "folderId": row["folder_id"], "version": row["version"],
                    "data": json.loads(row["data"])})


@app.route("/api/analyses/<aid>/folder", methods=["PATCH"])
def move_analysis_folder(aid):
    body = request.get_json() or {}
    folder_id = body.get("folderId")
    now = datetime.now().strftime("%b %d, %I:%M %p")
    conn = get_db()
    row = conn.execute("SELECT id FROM analyses WHERE id=?", (aid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    if folder_id is not None:
        frow = conn.execute("SELECT id FROM folders WHERE id=?", (folder_id,)).fetchone()
        if not frow:
            conn.close()
            return jsonify({"error": "Folder not found"}), 404
    conn.execute("UPDATE analyses SET folder_id=?, saved_at=? WHERE id=?", (folder_id, now, aid))
    conn.commit()
    conn.close()
    return jsonify({"id": aid, "folderId": folder_id, "savedAt": now})

@app.route("/api/analyses/<aid>", methods=["DELETE"])
def delete_analysis(aid):
    conn = get_db()
    conn.execute("DELETE FROM analyses WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": aid})


# ─────────────────────────────────────────────────────────────
# COLORS — only what the frontend actually uses
# ─────────────────────────────────────────────────────────────
# Backgrounds
CY   = "FFFFFBEB"  # yellow  — user input cells only
CW   = "FFFFFFFF"  # white   — all computed/data cells
CG8  = "FF1F2937"  # dark    — section headers, group header row bg
CG7  = "FF374151"  # mid-dark — KPI sub-label rows
CG1  = "FFF1F3F5"  # light   — totals row bg
CG50 = "FFF8F9FA"  # near-white — scenario input label cells

# Text colors
FW   = "FFFFFFFF"  # white (on dark bg)
FDK  = "FF111827"  # near-black — default data text
FG5  = "FF6B7280"  # grey — subtitles
FG4  = "FF9CA3AF"  # lighter grey
FYL  = "FFFCD34D"  # amber-brown — yellow input values & "VENDOR INPUTS" label
FGR  = "FF15803D"  # green — "RECIPE"/"COST" label + green values
FTEAL= "FF5EEAD4"  # teal — "COST" group label
FOR_ = "FFC2410C"  # orange — "MOQ REFERENCE" label
FPU  = "FF6D28D9"  # purple — "ANCHOR ANALYSIS" label
FRD  = "FFDC2626"  # red — "ELIG." label + negative/risk values
FBL  = "FF1D4ED8"  # blue — blue values (Min MOQ Cost card etc.)
FAMB = "FFD97706"  # amber — pct-sum highlight

# Number format strings
NF_INT   = "#,##0"
NF_DEC2  = "#,##0.00"
NF_DEC3  = "#,##0.000"
NF_PCT   = "0.00%"
NF_PCT1  = "0.0%"
NF_G1    = '#,##0.0" g"'
NF_G2    = '#,##0.00" g"'
NF_D2    = '"$"#,##0.00'
NF_D4    = '"$"#,##0.0000'
NF_D6    = '"$"#,##0.000000'
NF_UNITS = '#,##0 "units"'
# Delta: RED when positive (costs more = bad), GREEN when negative (costs less = good), dash for zero
NF_DELTA_G = '[Red]"+$"#,##0.000000;[Green]"-$"#,##0.000000;"-"'
NF_DELTA_U = '[Red]"+$"#,##0.0000;[Green]"-$"#,##0.0000;"-"'


def _fill(c):  return PatternFill("solid", fgColor=c)
def _font(bold=False, sz=11, color=FDK):
    return Font(name="Calibri", size=sz, bold=bold, color=color)
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _thin():
    s = Side(style="thin", color="FFD1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def _dark():
    s = Side(style="thin", color="FF374151")
    return Border(left=s, right=s, top=s, bottom=s)
def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _sec(ws, r, c1, c2, text, sz=11):
    """Dark full-width section bar."""
    if c1 != c2:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.fill = _fill(CG8); cell.font = _font(True, sz, FW)
    cell.alignment = _align("left"); cell.border = _dark()

def _grp(ws, r, c1, c2, text, fc=FW):
    """Group header spanning columns — dark bg, colored text label."""
    if c1 != c2:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.fill = _fill(CG8); cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
    cell.alignment = _align("center"); cell.border = _dark()

def _ch(ws, r, c, text, fc=FW):
    """Column header cell — dark bg."""
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = _fill(CG8); cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _dark()

def _inp(ws, r, c, val, al="right", nf=None):
    """Yellow user-input cell."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = _fill(CY); cell.font = _font(False, 11, FDK)
    cell.alignment = _align(al); cell.border = _thin()
    if nf: cell.number_format = nf

def _dat(ws, r, c, val, fc=FDK, bold=False, al="right", nf=None, sz=11):
    """White computed/data cell — no background tint."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = _fill(CW); cell.font = _font(bold, sz, fc)
    cell.alignment = _align(al); cell.border = _thin()
    if nf: cell.number_format = nf

def _tot(ws, r, c, val, fc=FDK, bold=True, al="right", nf=None):
    """Totals row cell — light grey bg."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = _fill(CG1); cell.font = _font(bold, 11, fc)
    cell.alignment = _align(al); cell.border = _thin()
    if nf: cell.number_format = nf


# ─────────────────────────────────────────────────────────────
# UNIT CONVERSION FORMULA FRAGMENTS
# ─────────────────────────────────────────────────────────────
def _u2g(val, unit_ref):
    return (f'({val}*IF({unit_ref}="lb",453.592,'
            f'IF({unit_ref}="kg",1000,'
            f'IF({unit_ref}="oz",28.34952,'
            f'IF({unit_ref}="gal",3785.41,1)))))')

def _g2u(grams, unit_ref):
    return (f'({grams}/IF({unit_ref}="lb",453.592,'
            f'IF({unit_ref}="kg",1000,'
            f'IF({unit_ref}="oz",28.34952,'
            f'IF({unit_ref}="gal",3785.41,1)))))')


# ═════════════════════════════════════════════════════════════
# COLUMN MAP  (18 columns, matching frontend exactly)
#
#  A(1)  Ingredient name           — user input (yellow, left)
#  B(2)  MOQ amount                — user input (yellow)
#  C(3)  MOQ unit                  — user input (yellow, center)
#  D(4)  PI amount                 — user input (yellow)
#  E(5)  PI unit                   — user input (yellow, center)
#  F(6)  Cost/unit amount          — user input (yellow)
#  G(7)  Cost unit                 — user input (yellow, center)
#  H(8)  Formula %                 — user input (yellow)
#  I(9)  g/Unit                    — computed  (white)
#  J(10) Cost/g                    — computed  (white)
#  K(11) Cost/Unit                 — computed  (white)
#  L(12) MOQ (g)                   — computed  (white)
#  M(13) PI (g)                    — computed  (white)
#  N(14) MOQ Cost $                — computed  (white)
#  O(15) Anchor Score              — computed  (white)
#  P(16) Units to Hit              — computed  (white)
#  Q(17) Rank                      — computed  (white, center)
#  R(18) Elig Y/N                  — user input (yellow, center)
# ═════════════════════════════════════════════════════════════

def build_excel(analysis_name, unit_weight_g, root_sku, anchor_util_pct, what_if_units, ingredients):
    wb = Workbook()
    ws = wb.active
    ws.title = "RM Analysis"
    N = len(ingredients)

    # Column widths — match the proportions visible in screenshot
    for col, width in {
        1: 24,   # Ingredient name
        2: 9,    # MOQ amount
        3: 5,    # MOQ unit
        4: 7,    # PI amount
        5: 5,    # PI unit
        6: 10,   # Cost/unit amount
        7: 5,    # Cost unit
        8: 10,   # Formula %
        9: 10,   # g/Unit
        10: 10,  # Cost/g
        11: 11,  # Cost/Unit
        12: 13,  # MOQ (g)
        13: 11,  # PI (g)
        14: 12,  # MOQ Cost $
        15: 12,  # Anchor Score
        16: 13,  # Units to Hit
        17: 7,   # Rank
        18: 6,   # Elig
    }.items():
        _w(ws, col, width)

    # ── Row index constants ───────────────────────────────────
    R_TITLE = 1
    R_LEGEND= 2
    R_S1    = 4
    R_SG    = 5
    R_ROOT  = 6
    R_AU    = 7
    R_S2    = 9
    R_KH    = 9
    R_KV    = 10
    R_KS    = 11
    R_GH    = 12   # group headers (VENDOR INPUTS / RECIPE / COST / MOQ REF / ANCHOR)
    R_CH    = 13   # column headers (MOQ / PI / COST/UNIT / FORMULA% / g/UNIT …)
    R_IF    = 14   # first ingredient row
    R_IL    = R_IF + N - 1
    R_IT    = R_IL + 1   # matrix totals
    R_S3    = R_IT + 2
    R_WH    = R_S3 + 1
    R_WI    = R_WH + 1
    R_CH2   = R_WI + 2
    R_CV    = R_CH2 + 1
    R_CI    = R_CV + 1
    R_CD    = R_CI + 1
    R_SH    = R_CD + 2
    R_SF    = R_SH + 1
    R_BH    = R_SF + 6
    R_BGH   = R_BH + 1
    R_BCH   = R_BGH + 1
    R_BF    = R_BCH + 1
    R_BL    = R_BF + N - 1
    R_BT    = R_BL + 1

    fr, lr  = R_IF, R_IL
    bf, bl  = R_BF, R_BL

    SG      = f"$B${R_SG}"
    AU      = f"$B${R_AU}"
    WI_REF  = f"$B${R_WI}"
    REC_MOQ = f"$A${R_KV}"
    plan_u  = f"IF({WI_REF}>0,{WI_REF},{REC_MOQ})"

    K_RNG   = f"$K${fr}:$K${lr}"
    O_RNG   = f"$O${fr}:$O${lr}"
    P_RNG   = f"$P${fr}:$P${lr}"
    Q_RNG   = f"$Q${fr}:$Q${lr}"
    EL_RNG  = f"$R${fr}:$R${lr}"

    # ── TITLE ─────────────────────────────────────────────────
    ws.row_dimensions[R_TITLE].height = 22
    _sec(ws, R_TITLE, 1, 18,
         f"  MOQ & Purchase Analysis Tool  ·  {analysis_name or 'Analysis'}", sz=13)

    # ── LEGEND row — yellow spans full width ──────────────────
    ws.row_dimensions[R_LEGEND].height = 14
    ws.merge_cells(start_row=R_LEGEND, start_column=1, end_row=R_LEGEND, end_column=18)
    lc = ws.cell(row=R_LEGEND, column=1, value="  🟡  Yellow = User Input")
    lc.fill = _fill(CY)
    lc.font = Font(name="Calibri", size=9, bold=True, color=FYL)
    lc.alignment = _align("left")
    lc.border = _dark()

    # ── SECTION 1 ─────────────────────────────────────────────
    ws.row_dimensions[R_S1].height = 16
    _sec(ws, R_S1, 1, 18, "  SECTION 1 — Scenario Parameters")

    for row, label, value, nf in [
        (R_SG, "Unit Weight (g)",         unit_weight_g,        NF_DEC2),
        (R_ROOT, "Root SKU",              root_sku,              None),
        (R_AU, "Anchor Utilization (%)",  anchor_util_pct/100, NF_PCT),
    ]:
        ws.row_dimensions[row].height = 18
        # label cell
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = _fill(CG50); lc.font = _font(True, 10, FDK)
        lc.alignment = _align("left"); lc.border = _thin()
        # value cell (yellow input)
        _inp(ws, row, 2, value, al="right", nf=nf)
        # rest of row — plain white
        for c in range(3, 19):
            cell = ws.cell(row=row, column=c)
            cell.fill = _fill(CW); cell.border = _thin()

    # ── SECTION 2 ─────────────────────────────────────────────
    ws.row_dimensions[R_S2].height = 16
    _sec(ws, R_S2, 1, 18, "  SECTION 2 — Recipe & Vendor Input  ·  Ingredient Matrix")

    # KPI cards
    elig_ct = f'COUNTIF({EL_RNG},"Y")'
    # Rec MOQ: weighted avg of top-3 UTH, weight = UTH × Cost/Unit (composite score)
    t3_num  = (f'SUMPRODUCT(({EL_RNG}="Y")*({Q_RNG}<=3)*({P_RNG}*{K_RNG}>0)*{P_RNG}*{P_RNG}*{K_RNG})')
    t3_den  = (f'SUMPRODUCT(({EL_RNG}="Y")*({Q_RNG}<=3)*({P_RNG}*{K_RNG}>0)*{P_RNG}*{K_RNG})')
    fp_num  = f'SUMPRODUCT(({EL_RNG}="Y")*({P_RNG}*{K_RNG}>0)*{P_RNG}*{P_RNG}*{K_RNG})'
    fp_den  = f'SUMPRODUCT(({EL_RNG}="Y")*({P_RNG}*{K_RNG}>0)*{P_RNG}*{K_RNG})'
    rec_f   = (f'=IF({elig_ct}>=3,'
               f'IFERROR(CEILING({t3_num}/{t3_den},1),0),'
               f'IFERROR(CEILING({fp_num}/{fp_den},1),0))')

    ws.row_dimensions[R_KH].height = 13
    ws.row_dimensions[R_KV].height = 30
    ws.row_dimensions[R_KS].height = 13

    kpi_cards = [
        # (c1, c2, label_text, formula, nf, value_fc, sub_text)
        (1,  4,  "RECOMMENDED MOQ",
         rec_f, NF_UNITS, FYL,
         "top 3 anchor weighted  (full-pool fallback if < 3 elig.)"),
        (5,  9,  "RECIPE COST / UNIT",
         f"=SUM(K{fr}:K{lr})", NF_D4, FGR,
         "at 100% efficiency  (ideal)"),
        (10, 13, "RECIPE COST / GRAM",
         f"=IFERROR(SUM(K{fr}:K{lr})/{SG},0)", NF_D6, FGR,
         "at 100% efficiency  (ideal)"),
        (14, 18, "MIN. MOQ COST",
         f"=SUM(N{fr}:N{lr})", NF_D2, FBL,
         "sum of all 1× MOQ purchase costs"),
    ]
    for c1, c2, lbl, formula, nf, fc, sub in kpi_cards:
        # header label
        ws.merge_cells(start_row=R_KH, start_column=c1, end_row=R_KH, end_column=c2)
        kh = ws.cell(row=R_KH, column=c1, value=lbl)
        kh.fill = _fill(CG7); kh.font = Font(name="Calibri", size=9, bold=True, color=FG4)
        kh.alignment = _align("left"); kh.border = _dark()
        # big value
        ws.merge_cells(start_row=R_KV, start_column=c1, end_row=R_KV, end_column=c2)
        kv = ws.cell(row=R_KV, column=c1, value=formula)
        kv.fill = _fill(CW); kv.font = Font(name="Calibri", size=16, bold=True, color=fc)
        kv.alignment = _align("left"); kv.border = _thin(); kv.number_format = nf
        # subtitle
        ws.merge_cells(start_row=R_KS, start_column=c1, end_row=R_KS, end_column=c2)
        ks = ws.cell(row=R_KS, column=c1, value=sub)
        ks.fill = _fill(CW); ks.font = Font(name="Calibri", size=9, color=FG5)
        ks.alignment = _align("left"); ks.border = _thin()

    # ── GROUP HEADERS (row R_GH) ──────────────────────────────
    # Matches frontend exactly:
    # Ingredient | VENDOR INPUTS(cols 2-8) | RECIPE(9) | COST(10-11) |
    # MOQ REFERENCE(12-14) | ANCHOR ANALYSIS(15-17) | ELIG.(18)
    ws.row_dimensions[R_GH].height = 14
    # Ingredient — spans rows R_GH and R_CH (rowspan=2 equivalent via merge)
    ws.merge_cells(start_row=R_GH, start_column=1, end_row=R_CH, end_column=1)
    ic = ws.cell(row=R_GH, column=1, value="INGREDIENT")
    ic.fill = _fill(CG8); ic.font = Font(name="Calibri", size=9, bold=True, color=FW)
    ic.alignment = Alignment(horizontal="left", vertical="center")
    ic.border = _dark()



    # ── COLUMN HEADERS (row R_CH) — sub-columns under each group ─
    # Under VENDOR INPUTS (2-8): MOQ | MOQ Unit | PI | PI Unit | Cost/unit | Cost Unit
    # Under RECIPE (9): Formula % | g/Unit   ← NOTE: frontend shows RECIPE spanning 2 cols
    # Re-check: screenshot shows RECIPE only over "Formula %" and then g/Unit is separate?
    # Looking at screenshot: Recipe group → Formula% + g/Unit (2 cols, 8-9)
    # Cost group → Cost/g + Cost/Unit (2 cols, 10-11)
    # So group headers should be:
    #   Ingredient(1) | Vendor Inputs(2-8=7cols) | Recipe(9=1col... wait
    # Screenshot: VENDOR INPUTS covers MOQ(num+unit=2) PI(num+unit=2) Cost/unit(num+unit=2) = 6 cols
    # But Formula% is also yellow (user input) under Vendor Inputs? No — it's separate "Recipe"
    # Let me recount from screenshot column by column:
    # Col1=Ingredient, Col2=MOQ, Col3=MOQUnit, Col4=PI, Col5=PIUnit, Col6=Cost/unit, Col7=CostUnit
    # Col8=Formula%, Col9=g/Unit, Col10=Cost/g, Col11=Cost/Unit
    # Col12=MOQ(g), Col13=PI(g), Col14=MOQ Cost$
    # Col15=Anchor Score, Col16=Units to Hit, Col17=Rank, Col18=Elig
    # Group spans: VendorInputs=2-7(6cols), Recipe=8-9(2cols), Cost=10-11(2cols),
    #              MOQRef=12-14(3cols), Anchor=15-17(3cols)
    # But above I already set VendorInputs=2-8 (7 cols). Fix: VendorInputs=2-7, Recipe=8-9
    # Need to redo group row. Already merged above — recreate properly.

    ws.row_dimensions[R_CH].height = 32

    col_hdrs = [
        # (col, text, group_fc for tinting, width_hint)
        (2,  "MOQ",         "FFFCD34D"),
        (3,  "MOQ\nUnit",   "FFFCD34D"),
        (4,  "PI",          "FFFCD34D"),
        (5,  "PI\nUnit",    "FFFCD34D"),
        (6,  "Cost /\nUnit","FFFCD34D"),
        (7,  "Cost\nUnit",  "FFFCD34D"),
        (8,  "Formula %",   "FF86EFAC"),
        (9,  "g / Unit",    "FF86EFAC"),
        (10, "Cost / g",    "FF5EEAD4"),
        (11, "Cost / Unit", "FF5EEAD4"),
        (12, "MOQ (g)",     "FFFDBA74"),
        (13, "PI (g)",      "FFFDBA74"),
        (14, "MOQ Cost $",  "FFFDBA74"),
        (15, "Anchor\nScore","FFC4B5FD"),
        (16, "Units\nto Hit","FFC4B5FD"),
        (17, "Rank",        "FFC4B5FD"),
        (18, "Elig.\nY/N",  "FFFCA5A5"),
    ]
    for col, text, fc in col_hdrs:
        cell = ws.cell(row=R_CH, column=col, value=text)
        cell.fill = _fill(CG8)
        cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _dark()

    # Fix group header merges to correct spans now that we know layout:
    # Unmerge the wrong ones and re-merge correctly
    # We need to redo R_GH row properly — unmerge all and redo
    # openpyxl doesn't have unmerge-all-in-row easily, so we'll just overwrite
    # the cells (merged cells keep first cell value; we need to fix 2-8 → 2-7, add 8-9)
    # Since we already called _grp with wrong spans above, we need to NOT do that —
    # let me restructure: do group headers AFTER column headers so we can overwrite.
    # Actually openpyxl merge_cells overwrites prior merges for overlapping ranges.
    # Re-do group header row with correct spans:
    ws.merge_cells(start_row=R_GH, start_column=2,  end_row=R_GH, end_column=7)
    ws.merge_cells(start_row=R_GH, start_column=8,  end_row=R_GH, end_column=9)
    ws.merge_cells(start_row=R_GH, start_column=10, end_row=R_GH, end_column=11)
    ws.merge_cells(start_row=R_GH, start_column=12, end_row=R_GH, end_column=14)
    ws.merge_cells(start_row=R_GH, start_column=15, end_row=R_GH, end_column=17)
    # col 18 no merge needed

    def _grp2(c1, text, fc):
        cell = ws.cell(row=R_GH, column=c1, value=text)
        cell.fill = _fill(CG8)
        cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
        cell.alignment = _align("center")
        cell.border = _dark()

    _grp2(2,  "VENDOR INPUTS",   "FFFCD34D")
    _grp2(8,  "RECIPE",          "FF86EFAC")
    _grp2(10, "COST",            "FF5EEAD4")
    _grp2(12, "MOQ REFERENCE",   "FFFDBA74")
    _grp2(15, "ANCHOR ANALYSIS", "FFC4B5FD")
    cell18 = ws.cell(row=R_GH, column=18, value="ELIG.")
    cell18.fill = _fill(CG8)
    cell18.font = Font(name="Calibri", size=9, bold=True, color="FFFCA5A5")
    cell18.alignment = _align("center"); cell18.border = _dark()

    # ── INGREDIENT ROWS ───────────────────────────────────────
    for i, ing in enumerate(ingredients):
        row = R_IF + i
        ws.row_dimensions[row].height = 16

        moq_amt   = float(ing.get("moq")        or 0)
        moq_unit  = str(ing.get("unit")         or "lb").strip() or "lb"
        pi_amt    = float(ing.get("pi")         or 0)
        pi_unit_r = str(ing.get("piUnit")       or "").strip()
        pi_unit   = pi_unit_r if pi_unit_r else moq_unit
        cost_amt  = float(ing.get("costPerLb")  or 0)
        cost_unit = str(ing.get("costUnit")     or "lb").strip() or "lb"
        pct_dec   = float(ing.get("pct")        or 0)
        anch_ovr  = ing.get("anchorOvr")
        elig      = anch_ovr if anch_ovr in ("Y","N") else ("Y" if cost_amt > 0 else "N")

        def r_(l): return f"{l}{row}"
        B=r_("B"); C=r_("C"); D=r_("D"); E=r_("E")
        F=r_("F"); G=r_("G"); H=r_("H"); I=r_("I"); J=r_("J")
        K=r_("K"); L=r_("L"); M=r_("M"); N_=r_("N"); O_=r_("O")
        P_=r_("P"); Q_=r_("Q"); R_=r_("R")

        # Yellow user-input cells
        _inp(ws, row, 1, ing.get("name",""), al="left")
        _inp(ws, row, 2, moq_amt,  al="right", nf=NF_DEC2)
        _inp(ws, row, 3, moq_unit, al="center")
        _inp(ws, row, 4, pi_amt,   al="right", nf=NF_DEC2)
        _inp(ws, row, 5, pi_unit,  al="center")
        _inp(ws, row, 6, cost_amt, al="right", nf='"$"#,##0.0000')
        _inp(ws, row, 7, cost_unit, al="center")
        _inp(ws, row, 8, pct_dec,  al="right", nf=NF_PCT)
        _inp(ws, row, 18, elig,    al="center")

        # White computed cells
        # g/Unit = unit_weight_g × formula%
        _dat(ws, row, 9,  f"={SG}*{H}", nf=NF_DEC3)
        # Cost/g = cost_per_cost_unit / grams_per_cost_unit
        _dat(ws, row, 10, f"=IFERROR({F}/{_u2g('1',G)},0)", nf=NF_D6)
        # Cost/Unit = g/unit × cost/g
        _dat(ws, row, 11, f"={I}*{J}", nf=NF_D4)
        # MOQ (g)
        _dat(ws, row, 12, f"={_u2g(B,C)}", nf=NF_DEC2)
        # PI (g) — mirrors frontend piGrams()
        pi_g = (f'=IFERROR(IF(OR({D}="",{D}=0),"—",{_u2g(D,E)}),"—")')
        _dat(ws, row, 13, pi_g, nf=NF_DEC2)
        # MOQ Cost $
        _dat(ws, row, 14, f"={L}*{J}", nf=NF_D2)
        # Anchor Score
        _dat(ws, row, 15, f'=IF({R_}="Y",{N_}*{AU},0)', nf=NF_D2)
        # Units to Hit
        _dat(ws, row, 16, f'=IFERROR(IF({R_}="Y",({L}*{AU})/{I},0),0)', nf=NF_DEC3)
        # Rank
        # Composite rank score = UTH × Cost/Unit (P × K)
        rank_f = (f'=IFERROR(IF(AND({R_}="Y",{P_}>0,{K}>0),'
                  f'1+SUMPRODUCT(({EL_RNG}="Y")*({P_RNG}*{K_RNG}>{P_}*{K})),'
                  f'"—"),"—")')
        _dat(ws, row, 17, rank_f, al="center")

    # ── MATRIX TOTALS ROW ─────────────────────────────────────
    ws.row_dimensions[R_IT].height = 17
    # Merge label across input cols
    ws.merge_cells(start_row=R_IT, start_column=1, end_row=R_IT, end_column=7)
    tc = ws.cell(row=R_IT, column=1, value="TOTALS")
    tc.fill = _fill(CG1); tc.font = _font(True, 11, FDK)
    tc.alignment = _align("left"); tc.border = _thin()

    _tot(ws, R_IT, 8,  f"=SUM(H{fr}:H{lr})", fc=FGR,  nf=NF_PCT)   # pct sum — green if 100
    _tot(ws, R_IT, 9,  f"={SG}",              fc=FDK,  nf=NF_G2)    # unit weight grams
    _tot(ws, R_IT, 10, "",      nf=None)
    _tot(ws, R_IT, 11, f"=SUM(K{fr}:K{lr})", fc=FDK,  nf=NF_D4)    # cost/unit sum
    _tot(ws, R_IT, 12, "",      nf=None)
    _tot(ws, R_IT, 13, "",      nf=None)
    _tot(ws, R_IT, 14, f"=SUM(N{fr}:N{lr})", fc=FDK,  nf=NF_D2)    # MOQ cost sum
    for c in range(15, 19):
        _tot(ws, R_IT, c, "")

    # ── SECTION 3 ─────────────────────────────────────────────
    ws.row_dimensions[R_S3].height = 16
    _sec(ws, R_S3, 1, 18, "  SECTION 3 — What-If Scenario Analysis")

    ws.row_dimensions[R_WH].height = 13
    _sec(ws, R_WH, 1, 18, "  What-If Input", sz=10)

    ws.row_dimensions[R_WI].height = 20
    lc2 = ws.cell(row=R_WI, column=1, value="What-If Units")
    lc2.fill = _fill(CG50); lc2.font = _font(True, 10, FDK)
    lc2.alignment = _align("left"); lc2.border = _thin()
    _inp(ws, R_WI, 2, int(what_if_units) if what_if_units else 0, al="right", nf=NF_INT)
    ws.merge_cells(start_row=R_WI, start_column=3, end_row=R_WI, end_column=12)
    hint = ws.cell(row=R_WI, column=3,
                   value=f'=IF({WI_REF}>0,"","← 0 = use Rec. MOQ")')
    hint.fill = _fill(CW); hint.font = Font(name="Calibri", size=9, color=FG5)
    hint.alignment = _align("left"); hint.border = _thin()
    for c in range(13, 19):
        ws.cell(row=R_WI, column=c).fill = _fill(CW)
        ws.cell(row=R_WI, column=c).border = _thin()

    # Cost comparison — breakdown cols forward refs
    F_sum = f"SUM(F{bf}:F{bl})"
    G_sum = f"SUM(G{bf}:G{bl})"
    H_sum = f"SUM(H{bf}:H{bl})"
    K_sum = f"SUM(K{bf}:K{bl})"

    # Delta format: positive = red (Δ +$X), negative = green (Δ -$X)
    cpg_wi    = f"=IFERROR({G_sum}/({plan_u}*{SG}),0)"
    cpg_ideal = f"=IFERROR(SUM(K{fr}:K{lr})/{SG},0)"
    cpu_wi    = f"=IFERROR({G_sum}/{plan_u},0)"
    cpu_ideal = f"=SUM(K{fr}:K{lr})"
    cpg_delta = f"={cpg_wi[1:]}-({cpg_ideal[1:]})"
    cpu_delta = f"={cpu_wi[1:]}-({cpu_ideal[1:]})"

    for r in [R_CH2, R_CV, R_CI, R_CD]:
        ws.row_dimensions[r].height = 18

    # Cost comparison header
    ws.merge_cells(start_row=R_CH2, start_column=1, end_row=R_CH2, end_column=9)
    ws.merge_cells(start_row=R_CH2, start_column=10, end_row=R_CH2, end_column=18)
    for c1, label in [(1, "COST / GRAM"), (10, "COST / UNIT")]:
        ch = ws.cell(row=R_CH2, column=c1, value=label)
        ch.fill = _fill(CG7); ch.font = Font(name="Calibri", size=10, bold=True, color=FW)
        ch.alignment = _align("center"); ch.border = _dark()

    # Sub-labels: What-If / Ideal
    for c1, c2 in [(1,4),(5,9),(10,13),(14,18)]:
        ws.merge_cells(start_row=R_CV, start_column=c1, end_row=R_CV, end_column=c2)
        ws.merge_cells(start_row=R_CI, start_column=c1, end_row=R_CI, end_column=c2)
    for c1, label in [(1,"What-If"),(5,"Ideal"),(10,"What-If"),(14,"Ideal")]:
        cv = ws.cell(row=R_CV, column=c1, value=label)
        cv.fill = _fill(CW); cv.font = Font(name="Calibri", size=10, bold=True, color=FG5)
        cv.alignment = _align("center"); cv.border = _thin()

    # Big metric values
    cmp = [
        (1,  cpg_wi,    FDK, NF_D6),
        (5,  cpg_ideal, FGR, NF_D6),
        (10, cpu_wi,    FDK, NF_D4),
        (14, cpu_ideal, FGR, NF_D4),
    ]
    for c1, formula, fc, nf in cmp:
        cv2 = ws.cell(row=R_CI, column=c1, value=formula)
        cv2.fill = _fill(CW); cv2.font = Font(name="Calibri", size=14, bold=True, color=fc)
        cv2.alignment = _align("center"); cv2.border = _thin(); cv2.number_format = nf

    # Delta row — red when positive, green when negative
    ws.merge_cells(start_row=R_CD, start_column=1, end_row=R_CD, end_column=9)
    ws.merge_cells(start_row=R_CD, start_column=10, end_row=R_CD, end_column=18)
    for c1, formula, nf in [(1, cpg_delta, NF_DELTA_G), (10, cpu_delta, NF_DELTA_U)]:
        cd = ws.cell(row=R_CD, column=c1, value=formula)
        cd.fill = _fill(CW); cd.font = Font(name="Calibri", size=11, bold=True, color=FDK)
        cd.alignment = _align("center"); cd.border = _thin(); cd.number_format = nf

    # ── SUMMARY BLOCK ─────────────────────────────────────────
    ws.row_dimensions[R_SH].height = 13
    ws.merge_cells(start_row=R_SH, start_column=1, end_row=R_SH, end_column=9)
    ws.merge_cells(start_row=R_SH, start_column=10, end_row=R_SH, end_column=18)
    for c1, label in [(1,"  GRAMS ANALYSIS"),(10,"  $ ANALYSIS")]:
        sh = ws.cell(row=R_SH, column=c1, value=label)
        sh.fill = _fill(CG8); sh.font = Font(name="Calibri", size=10, bold=True, color=FW)
        sh.alignment = _align("left"); sh.border = _dark()

    lef_g = f"({F_sum})-({H_sum})"
    lef_d = f"({G_sum})-({K_sum})"

    summary_rows = [
        ("Total Purchased", f"={F_sum}",                          NF_G1, FBL,
         "Total Purchased",  f"={G_sum}",                         NF_D2, FBL),
        ("Total Used",       f"={H_sum}",                         NF_G1, FGR,
         "Total Used",        f"={K_sum}",                        NF_D2, FGR),
        ("Total Leftover",   f"={lef_g}",                         NF_G1, FOR_,
         "Total Leftover",    f"={lef_d}",                        NF_D2, FRD),
        ("Gram Util %",      f"=IFERROR({H_sum}/{F_sum},0)",      NF_PCT1, FGR,
         "$ Util %",          f"=IFERROR({K_sum}/{G_sum},0)",     NF_PCT1, FGR),
        ("Gram At Risk %",   f"=IFERROR(({lef_g})/{F_sum},0)",    NF_PCT1, FRD,
         "$ At Risk %",       f"=IFERROR(({lef_d})/{G_sum},0)",   NF_PCT1, FRD),
    ]
    for idx, (gl, gf, gnf, gfc, dl, df, dnf, dfc) in enumerate(summary_rows):
        r = R_SF + idx
        ws.row_dimensions[r].height = 18
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        sl = ws.cell(row=r, column=1, value=gl)
        sl.fill = _fill(CW); sl.font = _font(True, 10, FDK)
        sl.alignment = _align("left"); sl.border = _thin()
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        sv = ws.cell(row=r, column=5, value=gf)
        sv.fill = _fill(CW); sv.font = Font(name="Calibri", size=13, bold=True, color=gfc)
        sv.alignment = _align("right"); sv.border = _thin(); sv.number_format = gnf
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=13)
        dl2 = ws.cell(row=r, column=10, value=dl)
        dl2.fill = _fill(CW); dl2.font = _font(True, 10, FDK)
        dl2.alignment = _align("left"); dl2.border = _thin()
        ws.merge_cells(start_row=r, start_column=14, end_row=r, end_column=18)
        dv = ws.cell(row=r, column=14, value=df)
        dv.fill = _fill(CW); dv.font = Font(name="Calibri", size=13, bold=True, color=dfc)
        dv.alignment = _align("right"); dv.border = _thin(); dv.number_format = dnf

    # ── BREAKDOWN TABLE ───────────────────────────────────────
    ws.row_dimensions[R_BH].height = 17
    ws.merge_cells(start_row=R_BH, start_column=1, end_row=R_BH, end_column=10)
    bh = ws.cell(row=R_BH, column=1, value="  PURCHASE BREAKDOWN")
    bh.fill = _fill(CG8); bh.font = _font(True, 11, FW)
    bh.alignment = _align("left"); bh.border = _dark()
    ws.merge_cells(start_row=R_BH, start_column=11, end_row=R_BH, end_column=18)
    bn = ws.cell(row=R_BH, column=11,
                 value=f'="@ "&TEXT({plan_u},"#,##0")&" units"')
    bn.fill = _fill(CG8); bn.font = Font(name="Calibri", size=9, color=FG4)
    bn.alignment = _align("right"); bn.border = _dark()

    # Breakdown group headers
    ws.row_dimensions[R_BGH].height = 13
    bkd_grps = [
        (1,  1,  "INGREDIENT",        FW),
        (2,  2,  "g NEEDED",          FPU),
        (3,  7,  "PURCHASE ANALYSIS", FOR_),
        (8,  10, "GRAMS",             FBL),
        (11, 13, "$",                 FGR),
    ]
    for c1, c2, label, fc in bkd_grps:
        if c1 != c2:
            ws.merge_cells(start_row=R_BGH, start_column=c1, end_row=R_BGH, end_column=c2)
        cell = ws.cell(row=R_BGH, column=c1, value=label)
        cell.fill = _fill(CG8); cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
        cell.alignment = _align("center"); cell.border = _dark()

    # Breakdown column headers
    ws.row_dimensions[R_BCH].height = 32
    bkd_cols = [
        (1,  "Ingredient",   FW),
        (2,  "g Needed",     FPU),
        (3,  "# MOQ",        FOR_),
        (4,  "# PI",         FOR_),
        (5,  "Buy Qty",      FOR_),
        (6,  "g to Buy",     FOR_),
        (7,  "Purchased $",  FOR_),
        (8,  "Used g",       FBL),
        (9,  "Leftover g",   FBL),
        (10, "Util %",       FBL),
        (11, "Used $",       FGR),
        (12, "Leftover $",   FGR),
        (13, "$ Util %",     FGR),
    ]
    for col, label, fc in bkd_cols:
        cell = ws.cell(row=R_BCH, column=col, value=label)
        cell.fill = _fill(CG8); cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _dark()

    # ── BREAKDOWN DATA ROWS ───────────────────────────────────
    for i in range(N):
        ir  = R_IF + i
        row = R_BF + i
        ws.row_dimensions[row].height = 16

        mA = f"A{ir}"; mC = f"C{ir}"; mI = f"I{ir}"
        mJ = f"J{ir}"; mL = f"L{ir}"; mM = f"M{ir}"

        pi_safe  = f"IF(ISNUMBER({mM}),{mM},0)"
        g_needed = f"={mI}*{plan_u}"
        n_moq    = "=1"
        n_pi     = (f"=IFERROR(IF(OR({pi_safe}=0,{mL}>=B{row}),0,"
                    f"CEILING((B{row}-{mL})/{pi_safe},1)),0)")
        g_buy    = (f"=IF({pi_safe}=0,MAX(B{row},{mL}),"
                    f"{mL}+D{row}*{pi_safe})")
        buy_qty  = (f'=IFERROR(TEXT({_g2u(f"F{row}",mC)},"#,##0.####")&" "&{mC},"—")')
        purch_d  = f"=F{row}*{mJ}"
        used_g   = f"=B{row}"
        lef_g_r  = f"=F{row}-B{row}"
        gutil    = f"=IFERROR(B{row}/F{row},0)"
        used_d   = f"=B{row}*{mJ}"
        lef_d_r  = f"=G{row}-K{row}"
        dutil    = f"=IFERROR(K{row}/G{row},0)"

        _dat(ws, row, 1,  f"={mA}",  al="left",   fc=FDK)
        _dat(ws, row, 2,  g_needed,  al="right",  nf=NF_DEC2)
        _dat(ws, row, 3,  n_moq,     al="center", nf=NF_INT)
        _dat(ws, row, 4,  n_pi,      al="center", nf=NF_INT)
        # Buy Qty bold to match frontend fontWeight:700
        bq = ws.cell(row=row, column=5, value=buy_qty)
        bq.fill = _fill(CW); bq.font = Font(name="Calibri", size=11, bold=True, color=FDK)
        bq.alignment = _align("right"); bq.border = _thin()
        _dat(ws, row, 6,  g_buy,     al="right",  nf=NF_DEC2)
        _dat(ws, row, 7,  purch_d,   al="right",  nf=NF_D2)
        _dat(ws, row, 8,  used_g,    al="right",  nf=NF_DEC2)
        _dat(ws, row, 9,  lef_g_r,   al="right",  nf=NF_DEC2)
        _dat(ws, row, 10, gutil,     al="center", nf=NF_PCT1)
        _dat(ws, row, 11, used_d,    al="right",  nf=NF_D2)
        _dat(ws, row, 12, lef_d_r,   al="right",  nf=NF_D2)
        _dat(ws, row, 13, dutil,     al="center", nf=NF_PCT1)

    # ── BREAKDOWN TOTALS ──────────────────────────────────────
    ws.row_dimensions[R_BT].height = 17
    _tot(ws, R_BT, 1,  "TOTALS",                                          al="left")
    _tot(ws, R_BT, 2,  f"=SUM(B{bf}:B{bl})",                             nf=NF_DEC2)
    _tot(ws, R_BT, 3,  ""); _tot(ws, R_BT, 4, ""); _tot(ws, R_BT, 5, "")
    _tot(ws, R_BT, 6,  f"=SUM(F{bf}:F{bl})",                             nf=NF_DEC2)
    _tot(ws, R_BT, 7,  f"=SUM(G{bf}:G{bl})",                             nf=NF_D2)
    _tot(ws, R_BT, 8,  f"=SUM(H{bf}:H{bl})",                             nf=NF_DEC2)
    _tot(ws, R_BT, 9,  f"=SUM(I{bf}:I{bl})",                             nf=NF_DEC2)
    _tot(ws, R_BT, 10, f"=IFERROR(SUM(H{bf}:H{bl})/SUM(F{bf}:F{bl}),0)", nf=NF_PCT1)
    _tot(ws, R_BT, 11, f"=SUM(K{bf}:K{bl})",                             nf=NF_D2)
    _tot(ws, R_BT, 12, f"=SUM(L{bf}:L{bl})",                             nf=NF_D2)
    _tot(ws, R_BT, 13, f"=IFERROR(SUM(K{bf}:K{bl})/SUM(G{bf}:G{bl}),0)", nf=NF_PCT1)

    # ── FOOTER ────────────────────────────────────────────────
    foot = R_BT + 2
    ws.row_dimensions[foot].height = 13
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=18)
    fc_cell = ws.cell(row=foot, column=1,
        value='="MOQ & Purchase Analysis Tool  ·  Generated: "&TEXT(TODAY(),"MMM D, YYYY")')
    fc_cell.fill = _fill(CW)
    fc_cell.font = Font(name="Calibri", size=9, color=FG5)
    fc_cell.alignment = _align("center")
    fc_cell.border = _thin()

    return wb


# ─────────────────────────────────────────────────────────────
# EXPORT ENDPOINT
# ─────────────────────────────────────────────────────────────
@app.route("/export", methods=["POST"])
def export():
    data = request.get_json()
    raw  = data.get("anchorUtil") or data.get("dollarWt")
    aup  = 95.0 if raw is None else float(raw)
    if aup <= 1.0: aup *= 100.0

    wb = build_excel(
        analysis_name   = data.get("analysisName", "Analysis"),
        unit_weight_g   = float(data.get("unitWeightG", data.get("sachetGrams", 0)) or 0),
        root_sku        = data.get("rootSKU", ""),
        anchor_util_pct = aup,
        what_if_units   = float(data.get("whatIfUnits",  0) or 0),
        ingredients     = data.get("ings", []),
    )
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe = re.sub(r'[^\w\s\-]', '_', data.get("analysisName") or "Analysis")
    return send_file(buf, as_attachment=True,
                     download_name=f"{safe}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─────────────────────────────────────────────────────────────
# SPA CATCH-ALL
# ─────────────────────────────────────────────────────────────
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_spa(path):
    if not app.static_folder or not os.path.isdir(app.static_folder):
        abort(404)
    if path.startswith("api/") or path == "export":
        abort(404)
    dist = os.path.join(app.static_folder, path)
    if os.path.isfile(dist):
        return send_file(dist)
    index = os.path.join(app.static_folder, "index.html")
    if os.path.isfile(index):
        return send_file(index)
    abort(404)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, debug=True)
